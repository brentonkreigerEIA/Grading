import pandas as pd
from datetime import datetime as dt
import os
import glob
import numpy as np

import GradeScrape

# Grab the course number and name from the main script "GradeScrape.py" as strings
c = GradeScrape.course
n = GradeScrape.name
lab = GradeScrape.program
sem = GradeScrape.semester
if lab =='EIA_WV':
    course_label = 'WV' + c
elif 'BP' in c:
    course_label = c
else:
    course_label = 'BP' + c

# Set up paths (make sure available offline on drive stream)
path_grading_master = "/Volumes/GoogleDrive/Shared drives/Ambassador Program/Committee_BridgeEDU/Bridge Program Coordinator/BEDU Grading 2021-2022.xlsx"

# Get some time data
year = dt.utcnow().strftime('%Y')
date = dt.utcnow().strftime('%Y-%m-%d-')

# Now for working with the csv

    # Get all downloads files that are csv files
csv_list = glob.glob("/Users/brentonkreiger/Downloads/*.csv")

    # Sort by modified time (0ldest to Newest is default, this is now in Newest to oldest)
    # files = list(filter(os.path.isfile, glob.glob('~/Downloads' + "*")))
csv_list.sort(key=lambda x: os.path.getmtime(x), reverse = True)
most_recent = csv_list[0]

# Read file into a Pandas data frame, where we will manipulate in Pandas
# Most recent should be the course that was just done, so no need to verify name
BP_data = pd.read_csv(most_recent)

# Add a blank row for University
BP_data.insert(1, 'University', "")

# Overwrite the spreadsheet (for testing)
# f = open("/Users/brentonkreiger/Desktop/testingforpy.csv",'w') #open file in write mode
# BP_data.to_csv(f, header=True, index=None) #no header and no index (line numbers)
# f.close()

# Use vlookup (as merge and join in pandas) to find University
    #read in team matching
df0 = pd.read_excel(path_grading_master, sheet_name='Team Matching', index_col=None, header=None) #from google stream
df0.to_csv("/Users/brentonkreiger/Desktop/match.csv", header=None, index=None) # send to a csv
df1 = pd.read_csv("/Users/brentonkreiger/Desktop/match.csv") # import the csv as df1 for matching
os.remove("/Users/brentonkreiger/Desktop/match.csv") # remove the interim file
    #grab the columns you want here
    #just the 4 team matching columns from TeamMatching
right_student = df1[['Student ID','Email','Username','University']]
    #just the ID, blank Uni
left_student = BP_data[['Student ID','University']]

# Left join operation provides all the rows from 1st dataframe and
# matching rows from the 2nd dataframe. If the rows are not matched
# in the 2nd dataframe then they will be replaced by NaN. Now we will
# try to match df3 to df4 which has no university just student ID

left_join = pd.merge(left_student,
                      right_student,
                      on='Student ID', #match based on Student ID
                      how='left') #keeps the first dataframe and adds matching rows from second

# Extract the correct University order as a vector (note it is still named University_y)
Uni = left_join['University_y']

#Add this to the blank column
# f2 = pd.read_csv("/Users/brentonkreiger/Desktop/testingforpy.csv") #read in original as a data frame
BP_data.insert(1,'University Name',Uni) # add the matched University names
BP_data.drop('University', axis=1, inplace=True) # now drop the placeholder
# BP_final = open("/Users/brentonkreiger/Desktop/testingforpy.csv",'w') #open original file in write mode
#write BP_data, the edited file, to the sheet of the master file

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', -1)

#Pass through again and add missing values to a DF

    # df[df['column name'].isna()]
    # this gives us a DF of the missing people for each sheet
BP_lostboyz = BP_data[BP_data['University Name'].isna()]
BP_lostboyz = BP_lostboyz[['Student ID', 'University Name', 'Email', 'Username']]

    #append to a main frame
GradeScrape.lost_boyz = GradeScrape.lost_boyz.append(BP_lostboyz)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


append_df_to_excel(path_grading_master, BP_data, sheet_name=course_label, index=False, header=True, startrow=0)

if GradeScrape.i == len(GradeScrape.course_names):
    print(GradeScrape.lost_boyz)
    append_df_to_excel(path_grading_master, GradeScrape.lost_boyz, sheet_name='To Add', index=False, header=True, startrow=0)
